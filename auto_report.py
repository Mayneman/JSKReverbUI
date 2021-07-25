# Get values from run.csv
from decimal import Decimal

import pandas as pd
import win32com
from openpyxl import load_workbook
import os
from docxtpl import DocxTemplate, InlineImage
from PIL import ImageGrab
import datetime
import pythoncom
from docx.shared import Mm
import xlwings as xw
from shutil import copyfile
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))


# ---EXCEL---
def get_raw_values(filename):
    df = pd.read_csv(filename, engine='python')
    data_dict = {}
    for row in range(0, 18):
        rd = df.iloc[row]
        data_dict[df.iloc[row][0]] = [rd[1], rd[2], rd[3], rd[4], rd[5], rd[6], rd[7], rd[8]]
    data_dict['rh'] = df.iloc[19][1]
    data_dict['temp'] = df.iloc[19][2]
    data_dict['pressure'] = df.iloc[19][3]
    return data_dict


def update_excel(file):
    # open Excel app in the background
    pythoncom.CoInitialize()
    app_excel = xw.App(visible=False)
    print('app excel')
    wbk = xw.Book(file)
    print('book')
    wbk.api.RefreshAll()
    print('ref')
    wbk.save()
    print('save')
    # # kill Excel process
    app_excel.kill()
    print('kill')
    del app_excel
    print('del')
    return


# Empty Room
def to_excel(file, data):
    workbook = load_workbook(filename=file, read_only=False, keep_vba=True)
    print('loaded Workbook')
    ws = workbook.worksheets[0]
    list_of_rows = list(range(5,40,2))
    # Main Values
    for i in list_of_rows:
        freq = str(ws['A' + str(i)].value)
        entry = data[freq]
        ws['B' + str(i)] = float(entry[0])
        ws['C' + str(i)] = float(entry[1])
        ws['D' + str(i)] = float(entry[2])
        ws['E' + str(i)] = float(entry[3])
        ws['F' + str(i)] = float(entry[4])
        ws['G' + str(i)] = float(entry[5])
        ws['H' + str(i)] = float(entry[6])
        ws['I' + str(i)] = float(entry[7])
    print('Paste Values')
    # Env Variables
    ws['B43'] = data['rh']
    ws['C43'] = data['temp']
    ws['D43'] = data['pressure']
    workbook.save(file)
    print('Save Values')
    workbook.close()
    print('Close')
    return

def to_excel_sample(file, data):
    workbook = load_workbook(filename=file, read_only=False, keep_vba=True)
    print('loaded Workbook')
    ws = workbook.worksheets[1]
    list_of_rows = list(range(5,40,2))
    # Main Values
    for i in list_of_rows:
        freq = str(ws['D' + str(i)].value)
        entry = data[freq]
        ws['E' + str(i)] = float(entry[0])
        ws['F' + str(i)] = float(entry[1])
        ws['G' + str(i)] = float(entry[2])
        ws['H' + str(i)] = float(entry[3])
        ws['I' + str(i)] = float(entry[4])
        ws['J' + str(i)] = float(entry[5])
        ws['K' + str(i)] = float(entry[6])
        ws['L' + str(i)] = float(entry[7])
    print('Paste Values')
    # Env Variables
    ws['E43'] = data['rh']
    ws['F43'] = data['temp']
    ws['G43'] = data['pressure']
    workbook.save(file)
    print('Save Values (Sample)')
    workbook.close()
    print('Close')
    return


def save_excel(file, save):
    copyfile(file, save)
    return


# Get values for report
def get_excel(file):
    workbook = load_workbook(filename=file, data_only=True)
    ws = workbook.worksheets[0]
    value_range = list(range(20, 38))
    result_list = []
    for value in value_range:
        s = str(value)
        result_list.append((ws['L' + s].value, ws['M' + s].value))
    workbook.close()
    return result_list


# Removes unnecessary fields for ease of reading.
def reduce_excel(xlsm):
    workbook = load_workbook(filename=xlsm, data_only=True, keep_vba=True)
    # REMOVE 2,4,5,6
    del workbook['T2 With Sample']
    del workbook['Output for Report']
    del workbook['Absorption Area']
    del workbook['Background calcs']

    ws = workbook['Initial Parameters']
    # ZERO [9 TO 12 D]
    for cell in range(9, 13):
        ws['D' + str(cell)].value = None
    # ZERO [16 TO 33 D,F,H,J]
    cols = ['D', 'F', 'H', 'J']
    for row in range(16, 34):
        for col in cols:
            ws[col + str(row)].value = None
    workbook.save(xlsm)
    workbook.close()


def save_excel_image(inputExcelFilePath, outputPNGImagePath):
    # Open the excel application using win32com
    o = win32com.client.Dispatch("Excel.Application")
    # Disable alerts and visibility to the user
    o.Visible = 0
    o.DisplayAlerts = 0
    # Open workbook
    wb = o.Workbooks.Open(inputExcelFilePath)

    # Extract first sheet
    sheet = o.Sheets(4)
    print(sheet)
    for n, shape in enumerate(sheet.Shapes):
        print(shape)
        # Save shape to clipboard, then save what is in the clipboard to the file
        shape.Copy()
        image = ImageGrab.grabclipboard()
        # Saves the image into the existing png file (overwriting) TODO ***** Have try except?
        image.save(outputPNGImagePath, 'png')
        print('Image Saved to: ', outputPNGImagePath)
        pass
    pass
    wb.Close(True)
    o.Quit()
    return


# Complete entire csv data transaction.
def full_values(csv, xlsm, sample):
    file = ROOT_DIR + "\\ReportFiles\\rt_calc.xlsm"
    print(file)
    data = get_raw_values(csv)
    print('data')
    if sample:
        to_excel_sample(file, data)
    else:
        to_excel(file, data)
    print('to_excel')
    update_excel(file)
    print('update_excel')
    save_excel(file, xlsm)
    print('save')
    if not sample:
        reduce_excel(xlsm)
    values = get_excel(file)
    print('values')
    return values


def report_output(file):
    workbook = load_workbook(filename=file, data_only=True)
    ws = workbook.worksheets[3]
    hz_table = {}
    for value in list(range(16, 34)):
        s = str(value)
        hz_table[ws['A' + s].value] = {'t1': "%0.2f" % ws['B' + s].value, 't2': "%0.2f" % ws['C' + s].value,
                                       'oto': "%0.2f" % ws['D' + s].value}
    # Practical sound absorption coefficient
    psac = ["%0.2f" % ws['B39'].value, "%0.2f" % ws['C39'].value, "%0.2f" % ws['D39'].value, "%0.2f" % ws['E39'].value,
            "%0.2f" % ws['F39'].value, "%0.2f" % ws['G39'].value]
    # Weighted sound absorption coefficient
    wsac = ["%0.2f" % ws['C43'].value, ws['C44'].value, ws['C45'].value]
    # Single number ratings
    snr = ["%0.2f" % ws['B49'].value, "%0.2f" % ws['B50'].value]
    workbook.close()
    save_excel_image(file, ROOT_DIR + '\\ReportFiles\\chartImage.png')
    return hz_table, psac, wsac, snr


# ---WORD---
def changeValues(save_location, bracket_type, values):
    # Import Template
    template = DocxTemplate(ROOT_DIR + '\\ReportFiles\\Templates\\' + bracket_type + '.docx')
    # Date
    x = datetime.datetime.now()
    hz_table, psac, wsac, snr = report_output(ROOT_DIR + "\\ReportFiles\\rt_calc.xlsm")
    # TODO: Gather actual values.
    context = {
        'report_number': values[0],
        'issue_date': str(x.day) + '/' + str(x.month) + '/' + str(x.year),
        'test_date': str(x.day) + '/' + str(x.month) + '/' + str(x.year),
        'client': values[1],
        'specimen_name': values[2],
        'specimen_desc': values[3],
        'A': values[4],
        'B': values[5],
        'C': values[6],
        'temperature': values[7],
        'humidity': values[8],
        'pressure': values[9],
        'hz': hz_table,
        'psac': psac,
        'wsac': wsac,
        'snr': snr,
        'chart': InlineImage(template, ROOT_DIR + '\\ReportFiles\\chartImage.png', width=Mm(105))
    }
    print(context)
    # Apply Values
    template.render(context)
    # Save Template
    template.save(save_location + '\\Report.docx')

# TESTING BELOW CAN DELETE TO CLEAN CODE

# full_values(ROOT_DIR + "/ReportFiles/NO_SAMPLE.csv")
# update_excel(ROOT_DIR + "\\ReportFiles\\rt_calc.xlsm")
# values = get_excel(ROOT_DIR + "\\ReportFiles\\rt_calc.xlsm")
# print(values)
# save_excel_image('C:\\Users\\Lab PC\\Desktop\\lewis_test\\SAMPLE_CALCS.xlsm', ROOT_DIR + '\\ReportFiles\\chartImage.png')
# reduce_excel('C:\\Users\\Lab PC\\Desktop\\lewis_test\\NO_SAMPLE_CALCS.xlsm')